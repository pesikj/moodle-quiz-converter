import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

tree = ET.parse('questions.xml')
root = tree.getroot()

question_parsed = []
for question in root:
    question_data = {"options": []}
    if question.attrib["type"] == "multichoice":
        for child in question:
            if child.tag == "name":
                question_data["text"] = child[0].text
            if child.tag == "answer":
                correct = float(child.attrib["fraction"]) > 0
                text = child[0].text
                question_data["options"].append((text, correct))
        question_parsed.append(question_data)

wb = Workbook()
ws = wb.active
row = 1
colours = {True: Color(indexed=50), False: Color(indexed=53)}
for question in question_parsed:
    column = 1
    cell = ws.cell(row, column)
    cell.value = question["text"]
    for option in question["options"]:
        column += 1
        cell = ws.cell(row, column)
        cell.value = option[0]
        cell.fill = PatternFill("solid", fgColor=colours[option[1]])
    row += 1
wb.save("questions.xlsx")
